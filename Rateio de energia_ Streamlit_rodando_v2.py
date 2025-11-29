# ===================== IMPORTAÇÕES =====================
import streamlit as st              # Framework para apps web simples em Python
import pandas as pd                 # Manipulação de dados tabulares
import plotly.express as px         # Gráficos interativos
import io                           # Buffer de memória para criar arquivos em tempo real
from datetime import datetime       # Data e hora
from zoneinfo import ZoneInfo       # Fuso horário (para horário local de Blumenau)
from openpyxl.utils import get_column_letter  # Ajuste automático de largura das colunas no Excel

# ===================== CONFIGURAÇÃO DA PÁGINA =====================
st.set_page_config(page_title="Rateio de Energia", page_icon="💡", layout="wide")
st.title("💡 Rateio de Energia - Quitinetes")

# ===================== ESTADO (SESSION_STATE) =====================
# Guardamos histórico, último resultado e resumo para persistirem após cliques
if "historico" not in st.session_state:
    st.session_state.historico = pd.DataFrame()
if "df_resultado" not in st.session_state:
    st.session_state.df_resultado = None
if "resumo_resultado" not in st.session_state:
    st.session_state.resumo_resultado = None
if "alertas_resultado" not in st.session_state:
    st.session_state.alertas_resultado = []
# Estados para importação do mês anterior
if "prev_map" not in st.session_state:
    st.session_state.prev_map = {}
if "import_resumo" not in st.session_state:
    st.session_state.import_resumo = None

# ===================== IMPORTAÇÃO DO MÊS ANTERIOR =====================
st.header("📅 Mês anterior (importar backup)")
arquivo = st.file_uploader("Carregue a planilha Excel do mês anterior", type=["xlsx"])

if arquivo is not None:
    try:
        xls = pd.ExcelFile(arquivo)

        # Lê aba Resumo
        resumo_imp = pd.read_excel(xls, sheet_name="Resumo")

        # Lê aba Rateio (ou primeira aba se não existir)
        try:
            rateio_imp = pd.read_excel(xls, sheet_name="Rateio")
        except Exception:
            abas = xls.sheet_names
            rateio_imp = pd.read_excel(xls, sheet_name=abas[0]) if abas else pd.DataFrame()

        # 🔧 Força cabeçalhos consistentes se não existirem
        if "Unidade" not in rateio_imp.columns:
            rateio_imp.columns = ["Unidade", "Consumo (kWh)", "Valor (R$)"]

        # Monta o mapeamento de leituras anteriores
        st.session_state.prev_map = dict(zip(rateio_imp["Unidade"], rateio_imp["Consumo (kWh)"]))

        # Guarda o resumo importado
        st.session_state.import_resumo = resumo_imp

        # Função para extrair valores do resumo
        def get_item(item):
            try:
                ser = resumo_imp.loc[resumo_imp["Item"] == item, "Valor"]
                return ser.values[0] if len(ser.values) else None
            except Exception:
                return None

        # Aplica valores do backup com segurança
        def aplicar_valor_seguro(chave_session, valor, opcoes_validas):
            if valor in opcoes_validas:
                st.session_state[chave_session] = valor

        aplicar_valor_seguro("bandeira_tarifaria", get_item("Bandeira por faixa"), ["Verde", "Amarela", "Vermelha 1", "Vermelha 2"])
        aplicar_valor_seguro("metodo_rateio", get_item("Método de rateio"), ["Proporcional ao total da fatura", "Faixas individuais"])
        aplicar_valor_seguro("fonte_consumo", get_item("Fonte do consumo total"), ["Leituras do prédio", "Soma das quitinetes"])

        # Mensagens de sucesso
        st.success("Backup importado! Leituras anteriores e configurações foram aplicadas quando possível.")
        st.write("Resumo do mês anterior:")
        st.dataframe(resumo_imp)
        st.write("Rateio do mês anterior (usado como leitura anterior):")
        st.dataframe(rateio_imp)

    except Exception as e:
        st.error("Erro ao importar backup. Verifique se a planilha está correta.")
        st.write(e)

# ===================== SIDEBAR: CONFIGURAÇÕES DE TARIFA =====================
st.sidebar.header("⚙️ Tarifas Celesc (R$/kWh com tributos)")
# TE: Tarifa de Energia | TUSD: Tarifa de Uso do Sistema de Distribuição
# Usamos duas faixas: até 150 kWh e acima de 150 kWh (modelo comum de residenciais)
tarifas = {
    "te_ate_150": st.sidebar.number_input("TE até 150 kWh", value=0.392200, format="%.6f"),
    "te_acima_150": st.sidebar.number_input("TE acima 150 kWh", value=0.415851, format="%.6f"),
    "tusd_ate_150": st.sidebar.number_input("TUSD até 150 kWh", value=0.455333, format="%.6f"),
    "tusd_acima_150": st.sidebar.number_input("TUSD acima 150 kWh", value=0.482660, format="%.6f"),
}

# COSIP: Contribuição para custeio de iluminação pública (valor fixo na fatura)
cosip = st.sidebar.number_input("COSIP (R$)", value=17.01, format="%.2f")
# ===================== BANDEIRA TARIFÁRIA =====================
st.sidebar.header("🚩 Bandeira tarifária")
bandeira_sel = st.sidebar.radio(
    "Selecione a bandeira",
    ["Verde", "Amarela", "Vermelha 1", "Vermelha 2"],
    index=2,  # seleciona "Vermelha 1" como inicial
    key="bandeira_tarifaria"
)
usar_bandeira_por_faixa = st.sidebar.checkbox("Usar bandeira por faixa (como na fatura)", value=True)

# Valor único por bandeira (aplicado quando NÃO usamos faixa)
bandeira_valor_unico = {
    "Verde": 0.000000,
    "Amarela": 0.018660,
    "Vermelha 1": 0.044630,
    "Vermelha 2": 0.075660,
}[bandeira_sel]

# Valores por faixa (aplicados quando usamos faixa)
bandeira_por_faixa = {
    "ate_150": st.sidebar.number_input("Bandeira até 150 kWh", value=0.054400, format="%.6f"),
    "acima_150": st.sidebar.number_input("Bandeira acima 150 kWh", value=0.057660, format="%.6f"),
}

# ===================== MÉTODO DE RATEIO E FONTE DO CONSUMO =====================
st.sidebar.header("📊 Método de rateio")
# - Faixas individuais: calcula cada unidade como se fosse uma fatura separada
# - Proporcional: distribui o total da fatura proporcional ao consumo de cada unidade
metodo_rateio = st.sidebar.radio("Escolha o método:", ["Proporcional ao total da fatura", "Faixas individuais"])

st.sidebar.header("📏 Fonte do consumo total")
# - Leituras do prédio: usa o medidor principal para consumo total
# - Soma das quitinetes: soma os consumos informados de cada unidade
fonte_consumo = st.sidebar.radio("Definir consumo total por:", ["Leituras do prédio", "Soma das quitinetes"])

# ===================== FUNÇÕES DE CÁLCULO =====================
def calcular_valor_base(consumo_kwh: float) -> float:
    """
    Calcula o custo base (TE + TUSD + Bandeira) para um dado consumo em kWh.
    Não inclui COSIP.
    Usa o modelo de duas faixas: até 150 kWh e acima de 150 kWh.
    """
    c1 = min(consumo_kwh, 150.0)        # parte até 150 kWh
    c2 = max(consumo_kwh - 150.0, 0.0)  # excedente acima de 150 kWh

    # Calcula TE e TUSD por faixa
    te = c1 * tarifas["te_ate_150"] + c2 * tarifas["te_acima_150"]
    tusd = c1 * tarifas["tusd_ate_150"] + c2 * tarifas["tusd_acima_150"]

    # Aplica bandeira por faixa ou valor único
    if usar_bandeira_por_faixa:
        bandeira = c1 * bandeira_por_faixa["ate_150"] + c2 * bandeira_por_faixa["acima_150"]
    else:
        bandeira = consumo_kwh * bandeira_valor_unico

    return round(te + tusd + bandeira, 2)
def calcular_fatura_total(consumo_total_kwh: float) -> tuple[float, float]:
    """
    Retorna (total_fatura, valor_base_sem_cosip).
    - valor_base: TE + TUSD + Bandeira
    - total_fatura: valor_base + COSIP
    """
    valor_base = calcular_valor_base(consumo_total_kwh)
    total = round(valor_base + cosip, 2)
    return total, valor_base

def adicionar_historico(nome_simulacao: str, df: pd.DataFrame, valor_total: float, consumo_total: float) -> None:
    """
    Adiciona a simulação atual ao histórico. Cada linha do df vira uma linha no histórico,
    com colunas extras: Identificação, Consumo Total, Valor Total.
    """
    linha = df.copy()
    linha["Identificação"] = nome_simulacao
    linha["Consumo Total"] = consumo_total
    linha["Valor Total"] = valor_total
    st.session_state.historico = pd.concat([st.session_state.historico, linha.reset_index()], ignore_index=True)

# ===================== INTERFACE PRINCIPAL =====================
# Leituras do prédio (medidor principal)
st.header("🔢 Leituras do prédio")
col1, col2 = st.columns(2)
with col1:
    leitura_predio_ant = st.number_input("Leitura anterior do prédio (kWh)", min_value=0, step=1)
with col2:
    leitura_predio_at = st.number_input("Leitura atual do prédio (kWh)", min_value=0, step=1)

# Identificação da simulação com data/hora local de Blumenau
hora_local = datetime.now(ZoneInfo("America/Sao_Paulo"))
nome_simulacao = st.text_input("Identificação da simulação", value=hora_local.strftime("%d/%m/%Y %H:%M"))

# Leituras das quitinetes (cada unidade)
st.header("🏠 Leituras das quitinetes")
n = st.slider("Número de quitinetes", 1, 5, value=1)
consumos_individuais = []
nomes_inquilinos = []

for i in range(n):
    with st.expander(f"Quitinete {i+1}", expanded=True):
        nome = st.text_input(f"Nome do inquilino Q{i+1}", key=f"nome_{i}")
        nome_final = nome.strip() if nome.strip() else f"Q{i+1}"
        nomes_inquilinos.append(nome_final)

        c1col, c2col = st.columns(2)
        with c1col:
            # Preenchimento automático a partir do backup importado
            leitura_ant_default = 0
            try:
                if st.session_state.prev_map and nome_final in st.session_state.prev_map:
                    leitura_ant_default = int(float(st.session_state.prev_map[nome_final]))
            except (ValueError, TypeError, KeyError):
                leitura_ant_default = 0

            ant = st.number_input("Leitura anterior (kWh)", min_value=0, step=1, value=leitura_ant_default, key=f"ant_{i}")

        with c2col:
            at = st.number_input("Leitura atual (kWh)", min_value=0, step=1, value=0, key=f"at_{i}")

        consumo = max(at - ant, 0)  # nunca deixa negativo
        consumos_individuais.append(float(consumo))

# ===================== CÁLCULO (AO CLICAR) =====================
if st.button("Calcular"):
    # 🔧 Salva leituras do prédio no session_state
    st.session_state["leitura_predio_ant"] = leitura_predio_ant
    st.session_state["leitura_predio_at"] = leitura_predio_at

    # Determina consumo total conforme fonte
    if fonte_consumo == "Leituras do prédio":
        consumo_total = float(max(leitura_predio_at - leitura_predio_ant, 0))
    else:
        consumo_total = float(sum(consumos_individuais))

    # Valor total da fatura e base (sem COSIP)
    valor_total, valor_base = calcular_fatura_total(consumo_total)

    # Calcula valores por unidade conforme método
    if metodo_rateio == "Faixas individuais":
        # Cada unidade calcula como se fosse uma fatura própria
        valores_individuais = [calcular_valor_base(c) for c in consumos_individuais]
    else:
        # Proporcional ao total da fatura (protege contra divisão por zero)
        valores_individuais = [
            round((c / consumo_total) * valor_total, 2) if consumo_total > 0 else 0.0
            for c in consumos_individuais
        ]

    # Monta DataFrame principal com resultados por unidade
    df = pd.DataFrame(
        {"Consumo (kWh)": consumos_individuais, "Valor (R$)": valores_individuais},
        index=[f"Quitinete {i+1} - {nomes_inquilinos[i]}" for i in range(n)]
    )

    # Calcula Áreas Comuns (diferença entre total e soma das unidades)
    soma_consumo_individual = float(sum(consumos_individuais))
    soma_valores_individuais = float(sum(valores_individuais))
    consumo_areas_comuns = round(consumo_total - soma_consumo_individual, 2)
    valor_areas_comuns = round(valor_total - soma_valores_individuais, 2)

    # Normaliza ruídos de arredondamento muito pequenos
    if abs(consumo_areas_comuns) < 0.01:
        consumo_areas_comuns = 0.0
    if abs(valor_areas_comuns) < 0.01:
        valor_areas_comuns = 0.0

    # Lista de alertas (avisos) para inconsistências
    alertas = []
    if consumo_areas_comuns < 0:
        alertas.append("Consumo das quitinetes excede o consumo total do prédio. Ajustei Áreas Comuns para 0 kWh.")
        consumo_areas_comuns = 0.0
    if valor_areas_comuns < 0:
        alertas.append("Soma dos valores individuais excede o total da fatura. Ajustei Áreas Comuns para R$ 0,00.")
        valor_areas_comuns = 0.0

    # Adiciona linha de Áreas Comuns se houver valor/consumo relevante
    if (consumo_areas_comuns != 0.0) or (valor_areas_comuns != 0.0):
        df.loc["Áreas Comuns"] = [consumo_areas_comuns, valor_areas_comuns]

    # Mensagens de status
    st.success(f"Consumo total do prédio: {consumo_total} kWh")
    st.success(f"Valor base (TE+TUSD+Bandeira): R$ {valor_base:.2f}")
    st.success(f"Valor total da fatura: R$ {valor_total:.2f}")
    for msg in alertas:
        st.warning(msg)

    # Salva resultado em session_state
    st.session_state.df_resultado = df
    st.session_state.alertas_resultado = alertas
    st.session_state.resumo_resultado = {
        "Identificação": nome_simulacao,
        "Consumo total (kWh)": consumo_total,
        "Valor base (R$)": valor_base,
        "COSIP (R$)": cosip,
        "Total fatura (R$)": valor_total,
        "Bandeira por faixa": "Sim" if usar_bandeira_por_faixa else "Não",
        "Método de rateio": metodo_rateio,
        "Fonte do consumo total": fonte_consumo,
        "Leitura do prédio (kWh)": st.session_state["leitura_predio_at"]  # ✅ Correção aplicada
    }

    # Adiciona ao histórico (cada unidade + possíveis Áreas Comuns)
    adicionar_historico(nome_simulacao, df, valor_total, consumo_total)
# ===================== EXIBIÇÃO PERSISTENTE DE RESULTADOS =====================
# Mostra tabela, gráfico e botão de exportar mesmo após outras interações
if st.session_state.df_resultado is not None:
    st.subheader("📊 Rateio detalhado")
    st.dataframe(st.session_state.df_resultado.style.format({"Valor (R$)": "R${:,.2f}"}))

    st.subheader("📈 Consumo por unidade")
    df_plot = st.session_state.df_resultado.reset_index().rename(columns={"index": "Unidade"})
    fig = px.bar(
        df_plot, x="Unidade", y="Consumo (kWh)",
        text="Consumo (kWh)", color="Unidade",
        labels={"Unidade": "Unidade", "Consumo (kWh)": "Consumo (kWh)"}
    )
    fig.update_traces(textposition="outside")
    st.plotly_chart(fig, use_container_width=True)

    for msg in st.session_state.alertas_resultado:
        st.warning(msg)
import io
import pandas as pd
from openpyxl.utils import get_column_letter

# Cria buffer de memória para gerar o arquivo Excel em tempo real
buffer = io.BytesIO()
wrote_any_sheet = False  # Flag para saber se alguma aba foi escrita

with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    try:
        # === ABA RATEIO ===
        df_export = st.session_state.df_resultado.copy()
        df_export.index.name = "Unidade"

        # Adiciona coluna de leitura atual (quitinetes)
        df_export["Leitura atual (kWh)"] = 0
        for i, unidade in enumerate(df_export.index):
            leitura_atual = st.session_state.get(f"at_{i}", 0)
            df_export.loc[unidade, "Leitura atual (kWh)"] = leitura_atual

        # Preenche leitura do prédio na linha correspondente
        leitura_predio_at = st.session_state.get("leitura_predio_at", None)
        if leitura_predio_at is not None:
            if "Áreas Comuns" in df_export.index:
                df_export.loc["Áreas Comuns", "Leitura atual (kWh)"] = leitura_predio_at
            elif "Prédio" in df_export.index:
                df_export.loc["Prédio", "Leitura atual (kWh)"] = leitura_predio_at
            elif "Total" in df_export.index:
                df_export.loc["Total", "Leitura atual (kWh)"] = leitura_predio_at

        # Exporta aba Rateio
        df_export.to_excel(writer, sheet_name="Rateio", index=True)
        wrote_any_sheet = True

        # Ajusta largura das colunas da aba Rateio
        ws = writer.sheets["Rateio"]
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row,
                                      min_col=1, max_col=ws.max_column):
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # === ABA RESUMO ===
        resumo_dict = st.session_state.get("resumo_resultado") or {}

        # Garante que a leitura do prédio esteja registrada no resumo
        resumo_dict["Leitura do prédio (kWh)"] = leitura_predio_at if leitura_predio_at is not None else 0

        # Exporta aba Resumo
        resumo = pd.DataFrame(list(resumo_dict.items()), columns=["Item", "Valor"])
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        wrote_any_sheet = True

        # Ajusta largura das colunas da aba Resumo
        ws = writer.sheets["Resumo"]
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row,
                                      min_col=1, max_col=ws.max_column):
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # === ABA HISTÓRICO ===
        historico_df = st.session_state.get("historico")
        if isinstance(historico_df, pd.DataFrame) and not historico_df.empty:
            historico_df.to_excel(writer, sheet_name="Histórico", index=False)
            wrote_any_sheet = True

            # Ajusta largura das colunas da aba Histórico
            ws = writer.sheets["Histórico"]
            for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row,
                                          min_col=1, max_col=ws.max_column):
                max_length = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

    except Exception as e:
        # Em caso de erro, cria aba de erro com a mensagem
        pd.DataFrame({"Erro": [str(e)]}).to_excel(writer, sheet_name="Erro", index=False)
        wrote_any_sheet = True

    # Se nenhuma aba foi escrita, cria uma aba padrão
    if not wrote_any_sheet:
        pd.DataFrame({"Info": ["Sem dados para exportar"]}).to_excel(writer, sheet_name="Resumo", index=False)

# Prepara nome do arquivo com base na identificação
nome_id = st.session_state.get("resumo_resultado", {}).get("Identificação", hora_local.strftime("%d-%m-%Y_%H-%M"))
# Finaliza e prepara botão de download
buffer.seek(0)

# ✅ Nome seguro para o arquivo
nome_id = st.session_state.get("resumo_resultado", {}).get("Identificação")
if not nome_id:
    nome_id = hora_local.strftime("%d-%m-%Y_%H-%M")

st.download_button(
    label="⬇️ Baixar relatório em Excel",
    data=buffer,
    file_name=f"rateio_{str(nome_id).replace('/', '-').replace(':', '-')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
# Botão de download no Streamlit
st.download_button(
    label="⬇️ Baixar relatório em Excel",
    data=buffer,
    file_name=f"rateio_{str(nome_id).replace('/', '-').replace(':', '-')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
# ===================== ABA HISTÓRICO (SIMPLIFICADA) =====================
st.header("📅 Histórico de Rateios")
if not st.session_state.historico.empty:
    st.dataframe(st.session_state.historico)

    st.divider()
    # Botão para zerar todo o histórico e começar do zero
    if st.button("🧹 Iniciar novo histórico"):
        st.session_state.historico = pd.DataFrame()
        st.success("Histórico apagado com sucesso. Pronto para uma nova simulação.")
else:
    st.info("Nenhum registro no histórico ainda. Faça um cálculo para começar.")

# -------------------------------
# EXPLICAÇÕES DISCRETAS (FIM DA PÁGINA)
# -------------------------------
st.divider()
st.caption("Orientações rápidas sobre configurações")

with st.expander("🚩 Bandeiras tarifárias", expanded=False):
    st.markdown("""
    As bandeiras tarifárias indicam custos extras na geração de energia:

    - **Verde** → sem acréscimo  
    - **Amarela** → pequeno acréscimo por kWh  
    - **Vermelha 1 e 2** → acréscimos maiores

    Se usar **por faixa**, aplica-se:
    - Até 150 kWh → valor reduzido  
    - Acima de 150 kWh → valor cheio
    """)

with st.expander("📊 Bandeira por faixa (como na fatura)", expanded=False):
    st.markdown("""
    Aplica valores diferentes conforme o consumo:

    - **Até 150 kWh:** usa o valor reduzido  
    - **Acima de 150 kWh:** usa o valor cheio

    Como calculamos:
    - Consumo é separado em duas partes: até 150 kWh e excedente.
    - Somamos: (até 150 × valor reduzido) + (excedente × valor cheio).
    """)

with st.expander("🧮 Método de rateio", expanded=False):
    st.markdown("""
    **Faixas individuais**
    - Calcula cada unidade como se tivesse sua própria fatura.
    - Mais justo para quem consome pouco.
    - Ideal quando cada unidade tem medidor próprio.

    **Proporcional ao total da fatura**
    - Divide o total do prédio proporcional ao consumo de cada unidade.
    - Reflete exatamente a fatura real.
    - Ideal quando há um único medidor.
    """)

with st.expander("📏 Fonte de consumo total", expanded=False):
    st.markdown("""
    **Leituras do prédio**
    - Usa o medidor principal do prédio.
    - Geralmente mais preciso.

    **Soma das quitinetes**
    - Soma os consumos individuais informados.
    - Útil quando não há leitura do prédio ou ela está indisponível.
    """)

st.caption("Estas explicações são referenciais e não substituem as regras oficiais da concessionária.")

# -------------------------------
# RODAPÉ
# -------------------------------
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: grey; font-size: 14px;'>"
    "Desenvolvido por <strong>Rafael Guimarães dos Santos</strong> — Todos os direitos reservados ©"
    "</div>",
    unsafe_allow_html=True
)
